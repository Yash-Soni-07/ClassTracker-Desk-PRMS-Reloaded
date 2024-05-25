from openpyxl import Workbook, load_workbook
def Update_In_Excel(batch_='', report_list=[], possibility_=1, date_=''):
    #Globalizing some variables
    global yes_HW_given, yes_T_taken, call_again, yes_num,Temporary_Storage, Test_Score, yes_num
    #Loading Excel File
    wb = load_workbook(f"C:/Class/{batch_}.xlsx")#Loading the workbook
    ws = wb.create_sheet(f"{date_}")
    #---Heading---#
    if possibility_==1:
        ws.append(['Name','Attendance'])
    elif possibility_==2:
        ws.append(['Name','Attendance','Homework', 'Test Score'])
    elif possibility_==3:
        ws.append(['Name','Attendance','Homework'])
    elif possibility_==4:
        ws.append(['Name', 'Attendance','Test Score'])

    for report in report_list:
        ws.append(report)

    wb.save(f"C:/Class/{batch_}.xlsx")
